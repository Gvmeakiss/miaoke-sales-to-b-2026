# -*- coding: utf-8 -*-
"""
导出工具 - 参考 refer/difference_analysis.py

生成汇总表及各类型明细 sheet：
- 汇总表：固定行顺序，含四大类及细分、小记、5.not test、仅订单/发货单/订单及发货单/发票、总计、DMS/OMS发票清单
- 各类型明细：按细分场景拆分的明细 sheet
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from aqpp_scenarios import build_aqpp_scenario_report
from export_schema import apply_export_schema
from legacy_summary import build_fy25_format_summary


SUMMARY_EXTRA_NT = {
    '仅订单': 'NT-28',
    '仅发货单': 'NT-29',
    '仅发运单': 'NT-29',
    '仅订单及发货单': 'NT-31',
    '仅订单及发运单': 'NT-31',
    '仅发票': 'NT-30',
}


def add_company_code(df):
    """建立展示用公司代码；缺失值单列为“公司代码缺失”，不丢弃记录。"""
    out = df.copy()
    source = next((column for column in ('公司代码', '发票-公司代码', '发票-销售组织', '销售组织') if column in out.columns), None)
    if source:
        company = out[source].astype('string').str.strip().str.replace(r'\.0$', '', regex=True)
        out['公司代码'] = company.fillna('公司代码缺失').replace('', '公司代码缺失')
    else:
        out['公司代码'] = '公司代码缺失'
    return out


def _group_summary(df, group_columns, amount_col, invoice_total_amount):
    """通用分组汇总，保证公司明细合计来自同一总体DataFrame。"""
    work = df.copy()
    work['_汇总金额'] = pd.to_numeric(work[amount_col], errors='coerce').fillna(0) if amount_col in work.columns else 0.0
    summary = work.groupby(group_columns, dropna=False, as_index=False).agg(
        记录数=(group_columns[0], 'size'),
        发票金额=('_汇总金额', 'sum'),
    )
    total_rows = len(work)
    summary['记录数占比'] = summary['记录数'].map(lambda value: _pct_str(value, total_rows))
    summary['发票金额占比'] = summary['发票金额'].map(
        lambda value: calculate_invoice_amount_share(value, invoice_total_amount)
    )
    return summary


def build_overall_summary(df, amount_col, invoice_total_amount):
    """按场景大类生成总体汇总并追加总计。"""
    category_col = 'AQPP分类' if 'AQPP分类' in df.columns else '场景分类'
    summary = _group_summary(df, [category_col], amount_col, invoice_total_amount)
    summary = summary.rename(columns={category_col: '场景分类'})
    total = pd.DataFrame([{
        '场景分类': '总计',
        '记录数': len(df),
        '发票金额': float(pd.to_numeric(df[amount_col], errors='coerce').fillna(0).sum()) if amount_col in df.columns else 0.0,
        '记录数占比': '100.0%',
        '发票金额占比': calculate_invoice_amount_share(
            pd.to_numeric(df[amount_col], errors='coerce').fillna(0).sum() if amount_col in df.columns else 0,
            invoice_total_amount,
        ),
    }])
    return pd.concat([summary, total], ignore_index=True)


def build_company_summary(df, amount_col, invoice_total_amount):
    """按公司代码汇总；公司代码缺失记录保留为独立分组。"""
    return _group_summary(df, ['公司代码'], amount_col, invoice_total_amount)


def build_company_scenario_summary(df, amount_col, invoice_total_amount):
    """按公司及场景编码展示场景分布。"""
    code_col = 'AQPP场景编码' if 'AQPP场景编码' in df.columns else '场景编码'
    category_col = 'AQPP分类' if 'AQPP分类' in df.columns else '场景分类'
    desc_col = 'AQPP场景描述' if 'AQPP场景描述' in df.columns else '场景描述'
    summary = _group_summary(
        df,
        ['公司代码', code_col, category_col, desc_col],
        amount_col,
        invoice_total_amount,
    )
    return summary.rename(columns={
        code_col: '场景编码',
        category_col: '场景分类',
        desc_col: '场景描述',
    })


def build_invoice_inventory_summary(invoice_inventory, amount_label):
    """按原始渠道发票类型一类一行汇总；冲销前置剔除以“部分”披露。"""
    columns = [
        '发票类型代码', '发票类型描述', '是否参与匹配',
        '清单行数', 'SAP发票数', '开票数量（基本单位）', '发票不含税金额', '发票金额', '发票金额占比',
    ]
    if invoice_inventory is None or invoice_inventory.empty:
        return pd.DataFrame(columns=columns)
    work = invoice_inventory.copy()
    amount_candidates = (
        ('实际金额（ZFN1）', '开票金额', '含税金额')
        if amount_label == '开票金额'
        else ('含税金额', 'SAP开票含税金额', '实际金额（ZFN1）')
    )
    amount_col = next((c for c in amount_candidates if c in work.columns), None)
    quantity_col = next(
        (c for c in ('开票数量（基本单位数量）', 'SAP开票基本数量', '开票数量') if c in work.columns),
        None,
    )
    invoice_no_col = next(
        (c for c in ('SAP发票号', 'SAP发票编号', '发票-SAP发票号') if c in work.columns),
        None,
    )
    defaults = {
        '发票类型代码规范': '未知', '发票类型描述规范': '未知/待确认',
        '发票类型业务分类': '未知/待确认', '发票类型处理方式': '仅保留待确认明细',
        '发票类型可参与匹配': False,
    }
    for column, default in defaults.items():
        if column not in work.columns:
            work[column] = default
    work['_发票金额'] = pd.to_numeric(work[amount_col], errors='coerce').fillna(0) if amount_col else 0.0
    work['_发票不含税金额'] = pd.to_numeric(work['无税金额'], errors='coerce').fillna(0) if '无税金额' in work.columns else 0.0
    work['_开票数量'] = pd.to_numeric(work[quantity_col], errors='coerce').fillna(0) if quantity_col else 0.0
    work['_发票号'] = work[invoice_no_col].astype('string') if invoice_no_col else pd.Series(pd.NA, index=work.index, dtype='string')
    group_columns = ['发票类型代码规范', '发票类型描述规范']
    summary = work.groupby(group_columns, dropna=False, as_index=False).agg(
        清单行数=('_发票金额', 'size'),
        SAP发票数=('_发票号', 'nunique'),
        _参与数量=('发票类型可参与匹配', 'sum'),
        **{
            '开票数量（基本单位）': ('_开票数量', 'sum'),
            '发票不含税金额': ('_发票不含税金额', 'sum'),
            '发票金额': ('_发票金额', 'sum'),
        },
    )
    summary = summary.rename(columns={
        '发票类型代码规范': '发票类型代码', '发票类型描述规范': '发票类型描述',
    })
    summary['是否参与匹配'] = '否'
    summary.loc[summary['_参与数量'].eq(summary['清单行数']), '是否参与匹配'] = '是'
    summary.loc[
        summary['_参与数量'].gt(0) & summary['_参与数量'].lt(summary['清单行数']),
        '是否参与匹配',
    ] = '部分（含冲销前置剔除）'
    inventory_total_amount = float(work['_发票金额'].sum())
    summary['发票金额占比'] = summary['发票金额'].map(
        lambda value: calculate_invoice_amount_share(value, inventory_total_amount)
    )
    total = pd.DataFrame([{
        '发票类型代码': '总计', '发票类型描述': '渠道发票清单总计',
        '是否参与匹配': '', '清单行数': len(work),
        'SAP发票数': int(work['_发票号'].nunique()),
        '开票数量（基本单位）': float(work['_开票数量'].sum()),
        '发票不含税金额': float(work['_发票不含税金额'].sum()),
        '发票金额': float(work['_发票金额'].sum()),
        '发票金额占比': calculate_invoice_amount_share(inventory_total_amount, inventory_total_amount),
    }])
    return pd.concat([summary[columns], total[columns]], ignore_index=True)


def build_invoice_scope_bridge(
    invoice_inventory,
    amount_label,
    aqpp_input_invoices=None,
    invalid_key_invoices=None,
):
    """桥接PBC至正式聚合输入，并将政策剔除与关键键缺失分开披露。"""
    columns = ['桥接项目', '清单行数', 'SAP发票数', '发票不含税金额', '发票金额']
    if invoice_inventory is None or invoice_inventory.empty:
        return pd.DataFrame(columns=columns)
    work = invoice_inventory.copy()
    amount_candidates = (
        ('实际金额（ZFN1）', '开票金额', '含税金额')
        if amount_label == '开票金额'
        else ('含税金额', 'SAP开票含税金额', '实际金额（ZFN1）')
    )
    amount_col = next((column for column in amount_candidates if column in work.columns), None)
    invoice_no_col = next(
        (column for column in ('SAP发票号', 'SAP发票编号', '发票-SAP发票号') if column in work.columns),
        None,
    )
    work['_桥接金额'] = pd.to_numeric(work[amount_col], errors='coerce').fillna(0) if amount_col else 0.0
    work['_桥接不含税金额'] = pd.to_numeric(work['无税金额'], errors='coerce').fillna(0) if '无税金额' in work.columns else 0.0
    work['_桥接发票号'] = work[invoice_no_col].astype('string') if invoice_no_col else pd.Series(pd.NA, index=work.index, dtype='string')
    matchable = work.get(
        '发票类型可参与匹配', pd.Series(False, index=work.index)
    ).fillna(False).astype(bool)

    def prepared(frame):
        if frame is None or frame.empty:
            return work.iloc[0:0].copy()
        result = frame.copy()
        frame_amount_col = next((column for column in amount_candidates if column in result.columns), None)
        frame_invoice_no_col = next(
            (column for column in ('SAP发票号', 'SAP发票编号', '发票-SAP发票号') if column in result.columns),
            None,
        )
        result['_桥接金额'] = pd.to_numeric(result[frame_amount_col], errors='coerce').fillna(0) if frame_amount_col else 0.0
        result['_桥接不含税金额'] = pd.to_numeric(result['无税金额'], errors='coerce').fillna(0) if '无税金额' in result.columns else 0.0
        result['_桥接发票号'] = result[frame_invoice_no_col].astype('string') if frame_invoice_no_col else pd.Series(pd.NA, index=result.index, dtype='string')
        return result

    def bridge_row(label, frame):
        frame = prepared(frame)
        return {
            '桥接项目': label,
            '清单行数': len(frame),
            'SAP发票数': int(frame['_桥接发票号'].nunique()),
            '发票不含税金额': round(float(frame['_桥接不含税金额'].sum()), 2),
            '发票金额': round(float(frame['_桥接金额'].sum()), 2),
        }

    policy_allowed = work.loc[matchable]
    formal_input = aqpp_input_invoices if aqpp_input_invoices is not None else policy_allowed
    invalid_keys = invalid_key_invoices if invalid_key_invoices is not None else work.iloc[0:0]
    excluded = work.loc[~matchable]
    raw_row = bridge_row('1. 原始渠道发票PBC', work)
    policy_row = bridge_row('2. 政策允许（冲销处理后）', policy_allowed)
    aqpp_row = bridge_row('3. 匹配键完整的正式聚合输入', formal_input)
    invalid_row = bridge_row('4. 政策允许但匹配键缺失', invalid_keys)
    excluded_row = bridge_row('5. 政策或冲销前置排除', excluded)
    check = {
        '桥接项目': '6. 校验差额（1-3-4-5）',
        '清单行数': raw_row['清单行数'] - aqpp_row['清单行数'] - invalid_row['清单行数'] - excluded_row['清单行数'],
        'SAP发票数': '',
        '发票不含税金额': round(raw_row['发票不含税金额'] - aqpp_row['发票不含税金额'] - invalid_row['发票不含税金额'] - excluded_row['发票不含税金额'], 2),
        '发票金额': round(raw_row['发票金额'] - aqpp_row['发票金额'] - invalid_row['发票金额'] - excluded_row['发票金额'], 2),
    }
    return pd.DataFrame(
        [raw_row, policy_row, aqpp_row, invalid_row, excluded_row, check],
        columns=columns,
    )


def _style_worksheet(ws, summary=False):
    """对大数据工作表做轻量格式化，不逐行遍历以免拖慢导出。"""
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 28
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)

    if summary:
        widths = [34, 14, 14, 34, 22, 20, 24, 24]
        for index, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(index)].width = width
        for index, cell in enumerate(ws[1], 1):
            header = str(cell.value or '')
            if index > len(widths):
                ws.column_dimensions[get_column_letter(index)].width = max(
                    16, min(30, len(header) * 2 + 2)
                )
            if any(word in header for word in ('金额', '数量', '净额', '差额')) and '占比' not in header:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, index).number_format = '#,##0.00'
            elif any(word in header for word in ('行数', '发票数', '记录数', '组数')) and '占比' not in header:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, index).number_format = '#,##0'
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, 1).value in {'小记', '总计', '6. 校验差额（1-3-4-5）'}:
                for cell in ws[row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill('solid', fgColor='D9EAF7')
        return

    for index, cell in enumerate(ws[1], 1):
        header = str(cell.value or '')
        if any(word in header for word in ['名称', '描述', '细分场景']):
            width = 30
        elif any(word in header for word in ['时间', '日期']):
            width = 20
        elif any(word in header for word in ['order-item', '订单号', '发货单号', '发票号']):
            width = 22
        else:
            width = max(12, min(20, len(header) * 2 + 2))
        ws.column_dimensions[get_column_letter(index)].width = width


def _write_detail_sheets(writer, frame, base_name, max_rows):
    """写入明细并在超过Excel行限制时自动拆sheet。"""
    if frame is None or frame.empty:
        return
    for part_no, start in enumerate(range(0, len(frame), max_rows), start=1):
        suffix = '' if part_no == 1 else f'_{part_no}'
        sheet_name = f'{base_name[:31-len(suffix)]}{suffix}'[:31]
        frame.iloc[start:start + max_rows].to_excel(
            writer, sheet_name=sheet_name, index=False, na_rep='N/A'
        )
        _style_worksheet(writer.sheets[sheet_name])

def _pct_str(count, total):
    """安全计算占比，避免除零"""
    if total == 0:
        return 'N/A'
    return f"{(count / total * 100):.1f}%"


def calculate_invoice_amount_share(amount, invoice_total_amount):
    """以对应渠道发票清单净额为唯一分母计算金额占比。

    分母包含红字及冲销，不得通过拼接“全部数据/大类/子类”等重叠集合计算。
    """
    denom = pd.to_numeric(pd.Series([invoice_total_amount]), errors='coerce').iloc[0]
    if pd.isna(denom) or abs(float(denom)) < 1e-12:
        return 'N/A'
    try:
        return f"{(float(amount) / float(denom) * 100):.1f}%"
    except (ZeroDivisionError, TypeError):
        return 'N/A'


def _write_section_block(ws, start_row, title, frame):
    """在同一汇总sheet中写入带标题的数据块，返回下一块起始行。"""
    ws.cell(start_row, 1, title)
    ws.cell(start_row, 1).font = Font(bold=True, color='1F4E78', size=12)
    if frame is None or frame.empty:
        ws.cell(start_row + 1, 1, '（无数据）')
        return start_row + 3
    for col_idx, column in enumerate(frame.columns, start=1):
        cell = ws.cell(start_row + 1, col_idx, column)
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.font = Font(color='FFFFFF', bold=True)
    for row_idx, row in enumerate(frame.itertuples(index=False), start=start_row + 2):
        for col_idx, value in enumerate(row, start=1):
            cell_value = '' if pd.isna(value) else value
            ws.cell(row_idx, col_idx, cell_value)
    return start_row + 2 + len(frame) + 2


def _style_aqpp_scenario_sheet(ws):
    """格式化参考样式的 AQPP 场景汇总表。"""
    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    header_fill = PatternFill('solid', fgColor='1F4E78')
    not_test_fill = PatternFill('solid', fgColor='FFF2CC')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    widths = {
        'A': 10, 'B': 12, 'C': 42, 'D': 16, 'E': 36,
        'F': 16, 'G': 24, 'H': 12, 'I': 36, 'J': 12,
        'K': 10, 'L': 16, 'M': 14, 'N': 18, 'O': 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(2, ws.max_row + 1):
        code = str(ws.cell(row, 2).value or '')
        if code.startswith('NT-'):
            for cell in ws[row]:
                cell.fill = not_test_fill
        for col in (12, 14):
            ws.cell(row, col).number_format = '#,##0.00'


def _style_fy25_summary_sheet(ws):
    """按去年汇总表的财务审计版式设置固定列宽、层级和数值格式。"""
    from openpyxl.styles import Alignment, Border, Side

    ws.freeze_panes = 'A2'
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f'A1:H{ws.max_row}'
    widths = {'A': 24, 'B': 38, 'C': 14, 'D': 12, 'E': 24, 'F': 18, 'G': 24, 'H': 24}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    header_fill = PatternFill('solid', fgColor='1F4E78')
    section_fill = PatternFill('solid', fgColor='D9D9D9')
    total_fill = PatternFill('solid', fgColor='D9EAF7')
    thin = Side(style='thin', color='7F7F7F')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws.row_dimensions[1].height = 28

    main_labels = {'1.完全匹配', '2.数量一致金额有差异', '3.金额一致数量有差异', '4.均有差异', '5. not test'}
    total_labels = {'小计', '总计'}
    for row in range(2, ws.max_row + 1):
        category = str(ws.cell(row, 1).value or '')
        subcategory = str(ws.cell(row, 2).value or '')
        ws.row_dimensions[row].height = 22
        for cell in ws[row]:
            cell.border = border
        if category in main_labels:
            for cell in ws[row]:
                cell.fill = section_fill
                cell.font = Font(bold=True)
        elif category in total_labels:
            for cell in ws[row]:
                cell.fill = total_fill
                cell.font = Font(color='FF0000', bold=True)
        elif category.endswith('发票清单'):
            for cell in ws[row]:
                cell.font = Font(bold=True)
        if subcategory:
            ws.cell(row, 2).alignment = Alignment(horizontal='left', indent=1)

        ws.cell(row, 3).number_format = '#,##0;[Red](#,##0);-'
        ws.cell(row, 4).number_format = '0.00%;[Red](0.00%);-'
        ws.cell(row, 5).number_format = '#,##0.00;[Red](#,##0.00);-'
        ws.cell(row, 6).number_format = '0.00%;[Red](0.00%);-'
        ws.cell(row, 7).number_format = '#,##0.00;[Red](#,##0.00);-'
        ws.cell(row, 8).number_format = '0.00%;[Red](0.00%);-'


def _write_combined_summary_sheet(
    writer,
    overall_summary,
    company_summary,
    company_scenario,
    aqpp_scenario_report=None,
):
    """将指定汇总块按顺序写入单个“汇总”sheet；AQPP场景汇总另写独立sheet。"""
    ws = writer.book.create_sheet('汇总', 0)
    ws.sheet_view.showGridLines = False
    row = 1
    sections = [
        ('一、总体汇总', overall_summary),
        ('二、各公司汇总', company_summary),
        ('三、各公司场景分布', company_scenario),
    ]
    for title, frame in sections:
        row = _write_section_block(ws, row, title, frame)
    for index in range(1, 12):
        ws.column_dimensions[get_column_letter(index)].width = 18
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['D'].width = 36

    if aqpp_scenario_report is not None and not aqpp_scenario_report.empty:
        aqpp_scenario_report.to_excel(writer, sheet_name='AQPP场景汇总', index=False)
        _style_aqpp_scenario_sheet(writer.sheets['AQPP场景汇总'])


def _resolve_export_paths(output_file, summary_file=None, detail_file=None, unmatched_file=None):
    """根据主输出路径推导汇总 / 明细 / 其他未匹配三个文件路径。"""
    output_path = Path(output_file)
    if summary_file is None:
        summary_file = output_path.with_name(output_path.stem.replace('明细', '汇总') + output_path.suffix)
        if Path(summary_file) == output_path or '汇总' not in Path(summary_file).stem:
            summary_file = output_path.with_name(
                output_path.stem.replace('明细', '').rstrip('-') + '汇总' + output_path.suffix
            )
    if detail_file is None:
        detail_file = output_path if '明细' in output_path.stem else output_path.with_name(
            output_path.stem + '明细' + output_path.suffix
        )
    detail_path = Path(detail_file)
    if unmatched_file is None:
        unmatched_file = detail_path.with_name(
            detail_path.stem + '-其他未匹配' + detail_path.suffix
        )
    return Path(summary_file), detail_path, Path(unmatched_file)


def _prepare_detail_frame(frame, drop_cols=None):
    """明细导出前统一加公司代码、删诊断列、套用中文字段规范。"""
    if frame is None or len(frame) == 0:
        return pd.DataFrame()
    export_df = add_company_code(frame)
    for column in drop_cols or []:
        if column in export_df.columns:
            export_df = export_df.drop(columns=[column])
    export_df = apply_export_schema(export_df)
    if '发票-DMS销售单号' in export_df.columns:
        cols = [c for c in export_df.columns if c != '发票-DMS销售单号']
        cols.append('发票-DMS销售单号')
        export_df = export_df[cols]
    return export_df


def _invoice_stats_values(invoice_stats):
    """兼容旧tuple，并统一返回清单行数、发票数、匹配键数及金额。"""
    if invoice_stats is None:
        return {'清单行数': 0, 'SAP发票数': 0, '匹配键数': 0, '发票金额': 0.0}
    if isinstance(invoice_stats, dict):
        return {
            '清单行数': int(invoice_stats.get('清单行数', 0)),
            'SAP发票数': int(invoice_stats.get('SAP发票数', 0)),
            '匹配键数': int(invoice_stats.get('匹配键数', 0)),
            '发票金额': float(invoice_stats.get('发票金额', 0.0)),
        }
    return {
        '清单行数': int(invoice_stats[0]),
        'SAP发票数': 0,
        '匹配键数': 0,
        '发票金额': float(invoice_stats[1]),
    }


def build_summary_scope(df_data, extra_categories, amount_col):
    """构建含完整标准NT场景的窄汇总底表，不复制主明细的无关宽字段。"""
    required = [
        '公司代码', '发票-公司代码', '发票-销售组织', '销售组织',
        'AQPP场景编码', 'AQPP场景描述', 'AQPP分类', 'AQPP可分类',
        amount_col, '发票不含税金额', 'SAP开票不含税金额',
        '订单-发票金额', 'SAP-DMS订单金额',
    ]
    base_columns = [column for column in dict.fromkeys(required) if column and column in df_data.columns]
    parts = [df_data[base_columns].copy()]
    consumed_names = set()
    for category_name, nt_code in SUMMARY_EXTRA_NT.items():
        if category_name in consumed_names:
            continue
        frame = (extra_categories or {}).get(category_name)
        if frame is None or frame.empty:
            continue
        # 同义名称只应取其中一个，防止调用方同时传入“发货/发运”两套别名。
        consumed_names.add(category_name)
        if category_name == '仅发货单':
            consumed_names.add('仅发运单')
        elif category_name == '仅发运单':
            consumed_names.add('仅发货单')
        elif category_name == '仅订单及发货单':
            consumed_names.add('仅订单及发运单')
        elif category_name == '仅订单及发运单':
            consumed_names.add('仅订单及发货单')

        part = pd.DataFrame(index=frame.index)
        for column in ('公司代码', '发票-公司代码', '发票-销售组织', '销售组织'):
            if column in frame.columns:
                part[column] = frame[column]
        amount_source = next(
            (column for column in (amount_col, '开票金额', 'SAP开票含税金额', '含税金额', '实际金额（ZFN1）') if column and column in frame.columns),
            None,
        )
        part[amount_col] = pd.to_numeric(frame[amount_source], errors='coerce') if amount_source else 0.0
        untaxed_source = next(
            (column for column in ('发票不含税金额', 'SAP开票不含税金额', '无税金额') if column in frame.columns),
            None,
        )
        if untaxed_source:
            target = 'SAP开票不含税金额' if amount_col == 'SAP开票含税金额' else '发票不含税金额'
            part[target] = pd.to_numeric(frame[untaxed_source], errors='coerce')
        part['AQPP场景编码'] = nt_code
        part['AQPP场景描述'] = {
            'NT-28': '仅订单', 'NT-29': '仅发运单', 'NT-30': '仅开票', 'NT-31': '仅订单及发运单'
        }[nt_code]
        part['AQPP分类'] = 'Not Test'
        part['AQPP可分类'] = False
        parts.append(part.reset_index(drop=True))
    return pd.concat(parts, ignore_index=True, sort=False)


def export_with_classification(df_data, output_file, file_label='',
                               amount_col=None, amount_label='开票金额',
                               order_inv_diff_col=None, inv_minus_order_col=None,
                               drop_cols=None,
                               extra_categories=None,
                               invoice_stats=None,
                               special_invoices=None,
                               invoice_inventory=None,
                               aqpp_input_invoices=None,
                               invalid_key_invoices=None,
                               invoice_matchability_summary=None,
                               cancellation_details=None,
                               cancellation_summary=None,
                               summary_file=None,
                               detail_file=None,
                               unmatched_file=None):
    """
    导出三份 Excel：

    1. 汇总：总体 / 各公司 / 各公司场景，并单列原始发票类型汇总与AQPP场景汇总
    2. 明细：仅「三单匹配」一个 sheet（AQPP 可分类、能执行三单匹配的记录）
    3. 其他未匹配：Not Test 一个 sheet，以及仅发货单、仅订单及发货单、特殊发票明细等
    """
    EXCEL_MAX = 1_048_575

    if len(df_data) == 0:
        print(f"  {file_label}: 无数据，跳过导出")
        return

    if amount_col is None:
        for col in df_data.columns:
            if '开票金额' in str(col) or 'SAP开票含税金额' in str(col):
                amount_col = col
                amount_label = '开票金额' if '开票' in str(col) and 'SAP' not in str(col) else 'SAP开票含税金额'
                break

    df_data = add_company_code(df_data)
    extra_categories = extra_categories or {}
    drop_cols = drop_cols or []

    stats_values = _invoice_stats_values(invoice_stats)
    invoice_total_amount = stats_values['发票金额'] if invoice_stats is not None else (
        float(pd.to_numeric(df_data[amount_col], errors='coerce').sum()) if amount_col and amount_col in df_data.columns else 0.0
    )
    df_summary_scope = add_company_code(build_summary_scope(df_data, extra_categories, amount_col))
    df_overall_summary = build_overall_summary(df_summary_scope, amount_col, invoice_total_amount)
    df_company_summary = build_company_summary(df_summary_scope, amount_col, invoice_total_amount)
    df_company_scenario = build_company_scenario_summary(df_summary_scope, amount_col, invoice_total_amount)
    df_invoice_inventory_summary = build_invoice_inventory_summary(
        invoice_inventory, amount_label
    )
    df_invoice_scope_bridge = build_invoice_scope_bridge(
        invoice_inventory,
        amount_label,
        aqpp_input_invoices=aqpp_input_invoices,
        invalid_key_invoices=invalid_key_invoices,
    )
    df_aqpp_scenario_report = build_aqpp_scenario_report(
        df_summary_scope,
        amount_col=amount_col,
        order_inv_diff_col=order_inv_diff_col,
        inv_minus_order_col=inv_minus_order_col,
        invoice_total_amount=invoice_total_amount,
        untaxed_amount_col=next(
            (column for column in ('发票不含税金额', 'SAP开票不含税金额') if column in df_summary_scope.columns),
            None,
        ),
    )
    df_fy25_format_summary = build_fy25_format_summary(
        df_data,
        amount_col=amount_col,
        amount_label=amount_label,
        invoice_total_amount=invoice_total_amount,
        order_inv_diff_col=order_inv_diff_col,
        inv_minus_order_col=inv_minus_order_col,
        extra_categories=extra_categories,
        invoice_stats=invoice_stats,
        invoice_stats_label=f'{file_label}发票清单' if file_label else '渠道发票清单',
    )
    summary_path, detail_path, unmatched_path = _resolve_export_paths(
        output_file, summary_file, detail_file, unmatched_file
    )
    for path in (summary_path, detail_path, unmatched_path):
        path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(summary_path, engine='openpyxl') as w:
        _write_combined_summary_sheet(
            w,
            df_overall_summary,
            df_company_summary,
            df_company_scenario,
            aqpp_scenario_report=df_aqpp_scenario_report,
        )
        df_invoice_inventory_summary.to_excel(w, sheet_name='发票类型汇总', index=False)
        _style_worksheet(w.sheets['发票类型汇总'], summary=True)
        df_fy25_format_summary.to_excel(w, sheet_name='去年格式汇总', index=False)
        _style_fy25_summary_sheet(w.sheets['去年格式汇总'])
        if invoice_matchability_summary is not None and not invoice_matchability_summary.empty:
            invoice_matchability_summary.to_excel(w, sheet_name='OMS类型匹配能力', index=False)
            _style_worksheet(w.sheets['OMS类型匹配能力'], summary=True)
        if not df_invoice_scope_bridge.empty:
            df_invoice_scope_bridge.to_excel(w, sheet_name='PBC-AQPP桥接', index=False)
            _style_worksheet(w.sheets['PBC-AQPP桥接'], summary=True)
        if cancellation_summary is not None and not cancellation_summary.empty:
            cancellation_summary.to_excel(w, sheet_name='冲销处理汇总', index=False)
            _style_worksheet(w.sheets['冲销处理汇总'], summary=True)

    # 明细：按内部“场景分类”拆sheet；导出字段中不再重复显示场景分类。
    if 'AQPP可分类' in df_data.columns:
        matched_mask = df_data['AQPP可分类'].fillna(False).astype(bool)
    else:
        matched_mask = df_data.get('AQPP分类', pd.Series('', index=df_data.index)).ne('Not Test')
    with pd.ExcelWriter(detail_path, engine='openpyxl') as w:
        matched_source = df_data.loc[matched_mask]
        if matched_source.empty:
            pd.DataFrame({'说明': ['无可三单匹配记录']}).to_excel(w, sheet_name='三单匹配', index=False)
            _style_worksheet(w.sheets['三单匹配'])
        else:
            category_column = 'AQPP分类' if 'AQPP分类' in matched_source.columns else '场景分类'
            category_order = ('完全匹配', '金额差异', '数量差异', '数量+金额差异')
            for category in category_order:
                category_frame = matched_source.loc[matched_source[category_column].eq(category)]
                if category_frame.empty:
                    continue
                prepared = _prepare_detail_frame(category_frame, drop_cols)
                _write_detail_sheets(w, prepared, category, EXCEL_MAX)

    # 其他未匹配：Not Test + 仅订单/发货等 outer-join + 特殊发票
    not_test_detail = _prepare_detail_frame(df_data.loc[~matched_mask], drop_cols)
    unmatched_sheets = []
    if not not_test_detail.empty:
        unmatched_sheets.append(('Not Test', not_test_detail))

    for extra_name, extra_df in extra_categories.items():
        prepared = _prepare_detail_frame(extra_df, drop_cols)
        if not prepared.empty:
            unmatched_sheets.append((extra_name, prepared))

    special_matched = df_data[df_data.get('特殊发票标记', pd.Series(False, index=df_data.index)).fillna(False)].copy()
    if not special_matched.empty and '冲销处理编码' in special_matched.columns:
        special_matched = special_matched.loc[special_matched['冲销处理编码'].isna()].copy()
    review_only = add_company_code(special_invoices) if special_invoices is not None and not special_invoices.empty else pd.DataFrame()
    # 已进入正式冲销明细的行不再在“特殊发票明细”重复出现。
    if not review_only.empty and '冲销处理编码' in review_only.columns:
        review_only = review_only.loc[review_only['冲销处理编码'].isna()].copy()
    if not special_matched.empty:
        special_matched['特殊发票明细来源'] = '参与匹配的特殊发票'
    if not review_only.empty:
        review_only['特殊发票明细来源'] = '未参与匹配的待确认类型'
    special_detail = pd.concat([special_matched, review_only], ignore_index=True, sort=False)
    prepared_special = _prepare_detail_frame(special_detail, drop_cols)
    if not prepared_special.empty:
        unmatched_sheets.append(('特殊发票明细', prepared_special))

    prepared_cancellation = _prepare_detail_frame(cancellation_details, drop_cols)
    if not prepared_cancellation.empty:
        unmatched_sheets.append(('冲销配对明细', prepared_cancellation))

    with pd.ExcelWriter(unmatched_path, engine='openpyxl') as w:
        if not unmatched_sheets:
            pd.DataFrame({'说明': ['无其他未匹配记录']}).to_excel(w, sheet_name='Not Test', index=False)
            _style_worksheet(w.sheets['Not Test'])
        else:
            for sheet_name, frame in unmatched_sheets:
                _write_detail_sheets(w, frame, sheet_name, EXCEL_MAX)

    print(f"  {file_label}: 汇总 -> {summary_path}")
    print(f"  {file_label}: 明细(三单匹配) -> {detail_path}")
    print(f"  {file_label}: 其他未匹配 -> {unmatched_path}")
